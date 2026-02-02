"""比較エンジン - CSV/Excel比較・JSON差分検出・画像類似度判定"""
# Copyright (c) 2025-2026 HarmonicInsight / FPT Consulting Japan. All rights reserved.
from __future__ import annotations

import csv
import json
import logging
from io import StringIO
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)


class Comparator:
    """テスト結果と期待値を比較する"""

    def compare_dict(
        self,
        expected: dict[str, Any],
        actual: dict[str, Any],
    ) -> list[str]:
        """辞書の差分を検出する"""
        differences: list[str] = []
        all_keys = set(expected.keys()) | set(actual.keys())

        for key in sorted(all_keys):
            if key not in expected:
                differences.append(f"予期しないキー: '{key}' = {actual[key]}")
            elif key not in actual:
                differences.append(f"欠落キー: '{key}' (期待値: {expected[key]})")
            elif expected[key] != actual[key]:
                differences.append(
                    f"値の不一致: '{key}' 期待={expected[key]}, 実際={actual[key]}"
                )

        return differences

    def compare_csv(
        self,
        expected_path: Path,
        actual_path: Path,
        key_columns: list[str] | None = None,
    ) -> list[str]:
        """CSVファイルを比較する"""
        differences: list[str] = []

        try:
            with open(expected_path, encoding="utf-8-sig") as f:
                expected_rows = list(csv.DictReader(f))
            with open(actual_path, encoding="utf-8-sig") as f:
                actual_rows = list(csv.DictReader(f))
        except Exception as e:
            return [f"CSV読み込みエラー: {e}"]

        if len(expected_rows) != len(actual_rows):
            differences.append(
                f"行数の不一致: 期待={len(expected_rows)}, 実際={len(actual_rows)}"
            )

        # 行単位の比較
        for i, (exp_row, act_row) in enumerate(
            zip(expected_rows, actual_rows)
        ):
            row_diffs = self.compare_dict(exp_row, act_row)
            for diff in row_diffs:
                differences.append(f"行{i + 1}: {diff}")

        return differences

    def compare_json(
        self,
        expected: Any,
        actual: Any,
        path: str = "$",
    ) -> list[str]:
        """JSON構造を再帰的に比較する"""
        differences: list[str] = []

        if type(expected) != type(actual):
            differences.append(
                f"{path}: 型の不一致 期待={type(expected).__name__}, "
                f"実際={type(actual).__name__}"
            )
            return differences

        if isinstance(expected, dict):
            all_keys = set(expected.keys()) | set(actual.keys())
            for key in sorted(all_keys):
                new_path = f"{path}.{key}"
                if key not in expected:
                    differences.append(f"{new_path}: 予期しないキー")
                elif key not in actual:
                    differences.append(f"{new_path}: 欠落")
                else:
                    differences.extend(
                        self.compare_json(expected[key], actual[key], new_path)
                    )
        elif isinstance(expected, list):
            if len(expected) != len(actual):
                differences.append(
                    f"{path}: 配列長の不一致 期待={len(expected)}, 実際={len(actual)}"
                )
            for i, (e, a) in enumerate(zip(expected, actual)):
                differences.extend(
                    self.compare_json(e, a, f"{path}[{i}]")
                )
        elif expected != actual:
            differences.append(
                f"{path}: 値の不一致 期待={expected}, 実際={actual}"
            )

        return differences

    def compare_excel(
        self,
        expected_path: Path,
        actual_path: Path,
        sheet_name: str | None = None,
    ) -> list[str]:
        """Excelファイルを比較する"""
        differences: list[str] = []

        try:
            from openpyxl import load_workbook

            wb_exp = load_workbook(str(expected_path), read_only=True)
            wb_act = load_workbook(str(actual_path), read_only=True)

            sheets = [sheet_name] if sheet_name else wb_exp.sheetnames

            for sheet in sheets:
                if sheet not in wb_act.sheetnames:
                    differences.append(f"シート欠落: '{sheet}'")
                    continue

                ws_exp = wb_exp[sheet]
                ws_act = wb_act[sheet]

                for row_idx, (exp_row, act_row) in enumerate(
                    zip(ws_exp.iter_rows(values_only=True), ws_act.iter_rows(values_only=True)),
                    start=1,
                ):
                    for col_idx, (exp_val, act_val) in enumerate(
                        zip(exp_row, act_row), start=1
                    ):
                        if exp_val != act_val:
                            differences.append(
                                f"[{sheet}] ({row_idx},{col_idx}): "
                                f"期待={exp_val}, 実際={act_val}"
                            )

            wb_exp.close()
            wb_act.close()

        except ImportError:
            differences.append("openpyxlが必要です: pip install openpyxl")
        except Exception as e:
            differences.append(f"Excel比較エラー: {e}")

        return differences
